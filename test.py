import os
import yaml
import argparse
import pandas as pd
from openpyxl import load_workbook  

import torch
from torch.utils.data import DataLoader
from evaluate import evaluate_rgb

from dataset.semi import SemiDataset
from model import Build_RGB_DinoV2, Build_RGB_HGT_DinoV2


# Argument parser
parser = argparse.ArgumentParser(description='Revisiting Weak-to-Strong Consistency in Semi-Supervised Semantic Segmentation')

parser.add_argument('--config', type=str, required=True)
parser.add_argument('--save-path', type=str, required=True)
parser.add_argument('--test-id-path', type=str, required=True)


def main():
    args = parser.parse_args()
    cfg = yaml.load(open(args.config, "r"), Loader=yaml.Loader)

    # ----- build model and load checkpoint ----- #    
    model, dataset, split, method, _ = args.save_path.split('/')

    if 'heightmatch' in method:
        model = Build_RGB_HGT_DinoV2(cfg['nclass']).cuda()
    else:
        model = Build_RGB_DinoV2(cfg['nclass']).cuda()

    # ------------- load checkpoint ----------- #       
    checkpoint_path = os.path.join(args.save_path, 'best.pth')
    print(f"loading {checkpoint_path}")
    checkpoint = torch.load(checkpoint_path, weights_only=True)
    state_dict = checkpoint['model']
    new_state_dict = {}
    for k, v in state_dict.items():
        new_key = k.replace("module.", "")
        new_state_dict[new_key] = v

    model.load_state_dict(new_state_dict)

    # ------------ build dataloader ----------- #   
    test_set = SemiDataset(cfg['dataset'], cfg['data_root'], 'val', \
                            cfg['class_bd'], cfg['class_bg'], id_path=args.test_id_path)
    test_loader = DataLoader(test_set, batch_size=64, pin_memory=True, 
                             num_workers=16, drop_last=False)

    # ----------- make predictions ---------- #        
    iou_cls, f1_cls, prec_cls, recall_cls = evaluate_rgb(model, test_loader, cfg)

    iou_buildings, f1_buildings, prec_buildings, recall_buildings \
        = iou_cls[1], f1_cls[1], prec_cls[1], recall_cls[1]

    # --------- save results to excel ------- #        
    data = {
        'Dataset': [dataset],
        'Split': [split],
        'Method': [method],
        'Precision (Buildings)': [round(prec_buildings, 2)],
        'Recall (Buildings)': [round(recall_buildings, 2)],
        'IoU (Buildings)': [round(iou_buildings, 2)],
        'F1 (Buildings)': [round(f1_buildings, 2)],
    }

    df = pd.DataFrame(data)
    excel_file = 'test_results.xlsx'

    # Check if the Excel file exists
    if os.path.exists(excel_file):
        book = load_workbook(excel_file)

        if 'Results' in book.sheetnames:
            sheet = book['Results']
            for row in df.values:
                sheet.append(row.tolist())  
        else:
            sheet = book.create_sheet('Results')
            df.to_excel(excel_file, index=False, header=True, sheet_name='Results')

        book.save(excel_file)
    else:
        df.to_excel(excel_file, index=False, header=True, sheet_name='Results')

    print(f"Results saved to {excel_file}")


if __name__ == '__main__':
    main()
